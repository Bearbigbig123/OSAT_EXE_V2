import os
import pandas as pd
import numpy as np
from PyQt6 import QtWidgets, QtCore, QtGui
import pickle

# Translation System
from translations import get_translator, tr

# Check if openpyxl package is installed
try:
    import openpyxl
except ImportError:
    openpyxl = None

class FormulaExplanationDialog(QtWidgets.QDialog):
    """公式說明對話框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("formula_explanation"))
        self.setMinimumSize(900, 600)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        
        # 標題
        title_label = QtWidgets.QLabel("<h2 style='text-align:center; color:#344CB7;'>📖 Formula Explanation / 公式說明</h2>")
        title_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # 滾動區域
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        content_widget = QtWidgets.QWidget()
        content_layout = QtWidgets.QHBoxLayout(content_widget)
        content_layout.setContentsMargins(10, 10, 10, 10)
        content_layout.setSpacing(15)
        
        # ===== 左側：英文版 =====
        left_widget = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_widget)
        left_layout.setSpacing(15)
        
        notice_label_en = QtWidgets.QLabel()
        notice_label_en.setWordWrap(True)
        notice_label_en.setTextFormat(QtCore.Qt.TextFormat.RichText)
        notice_label_en.setFont(QtGui.QFont("Arial", 10))
        notice_html_en = """
<div style='background-color:#f5f5f5; padding:12px; border-radius:6px; font-size:13px; margin-bottom:15px; font-family:Arial;'>
  <strong style='font-size:15px;'>⚠ Notice</strong>
  <p style='margin:8px 0;'>Table only shows items requiring attention (abnormal detection criteria):</p>
  <ul style='margin:8px 0 8px 20px; padding-left:0; line-height:1.8;'>
    <li><span style='color:#d9534f;'><strong>mean_matching_index ≥ 1</strong></span>: Mean not matched</li>
    <li><span style='color:#d9534f;'><strong>sigma_matching_index ≥ K</strong></span>: Sigma not matched</li>
    <li><span style='color:#8a6d3b;'><strong>Insufficient Data</strong></span>: Sample size less than 5</li>
  </ul>
</div>
        """
        notice_label_en.setText(notice_html_en)
        left_layout.addWidget(notice_label_en)
        
        formula_label_en = QtWidgets.QLabel()
        formula_label_en.setWordWrap(True)
        formula_label_en.setTextFormat(QtCore.Qt.TextFormat.RichText)
        formula_label_en.setFont(QtGui.QFont("Arial", 10))
        formula_html_en = """
<div style="background-color:#e8f4f8; padding:15px; border-radius:6px; font-size:13px; line-height:1.6; font-family:Arial; border: 2px solid #344CB7;">
  <strong style='font-size:15px;'>📘 Calculation Formula</strong>
  <table style="font-size:12px; margin-top:12px; font-family:Arial; width:100%;">
    <tr>
      <td style="vertical-align:top; padding:8px; padding-right:12px; width:35%;"><strong>Mean Matching Index:</strong></td>
      <td style="padding:8px;">
        <u>Two-group comparison:</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>|μ₁ − μ₂| / min(σ₁, σ₂)</code><br><br>
        <u>Multi-group comparison:</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>|μ − median(μ)| / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding:8px; padding-right:12px;"><strong>Sigma Matching Index:</strong></td>
      <td style="padding:8px;">
        <u>Two-group comparison:</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>σ / min(σ₁, σ₂)</code><br><br>
        <u>Multi-group comparison:</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>σ / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding:8px; padding-right:12px;"><strong>K Value:</strong></td>
      <td style="padding:8px;">
        <code style='background:#fff; padding:6px 10px; border-radius:3px; display:block; line-height:1.8;'>
          n = Sample size<br>
          n ≤ 4: No comparison<br>
          5 ≤ n ≤ 10: K = 1.73<br>
          11 ≤ n ≤ 120: K = 1.414<br>
          n > 120: K = 1.15
        </code>
      </td>
    </tr>
  </table>
  <div style="margin-top:15px; padding:12px; background:#fff; border-radius:4px; font-size:12px; color:#344CB7;">
    <strong>【Calculation Description for Filter Mode】</strong><br>
    <ul style="margin:8px 0 8px 20px; padding-left:0; line-height:1.8;">
      <li><strong>Mean/Std/Sample Size:</strong> For each matching_group, take data within "one month before the specified date", fill to specified number if insufficient (default 5, adjustable via UI), then calculate mean/std/count.</li>
      <li><strong>Median(sigma):</strong> For each matching_group, take data within "six months before the specified date", fill to specified number if insufficient (default 5), calculate std for each group, then take median.</li>
      <li><strong>Chart Display:</strong> Based on data "within one month (filled to specified number)".</li>
      <li><strong>Multiple Data Points:</strong> If multiple data points exist for the same group at the same time, all are included in calculation.</li>
    </ul>
    <span style="color:#8a6d3b; font-style:italic;">• All Data mode directly uses all data for group calculation without filling</span>
  </div>
</div>
        """
        formula_label_en.setText(formula_html_en)
        left_layout.addWidget(formula_label_en)
        left_layout.addStretch()
        
        # ===== 右側：中文版 =====
        right_widget = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_widget)
        right_layout.setSpacing(15)
        
        notice_label_zh = QtWidgets.QLabel()
        notice_label_zh.setWordWrap(True)
        notice_label_zh.setTextFormat(QtCore.Qt.TextFormat.RichText)
        notice_label_zh.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        notice_html_zh = """
<div style='background-color:#f5f5f5; padding:12px; border-radius:6px; font-size:13px; margin-bottom:15px; font-family:Microsoft JhengHei;'>
  <strong style='font-size:15px;'>⚠ 注意</strong>
  <p style='margin:8px 0;'>表格僅顯示需要關注的項目（異常檢測標準）：</p>
  <ul style='margin:8px 0 8px 20px; padding-left:0; line-height:1.8;'>
    <li><span style='color:#d9534f;'><strong>mean_matching_index ≥ 1</strong></span>：平均值不匹配</li>
    <li><span style='color:#d9534f;'><strong>sigma_matching_index ≥ K</strong></span>：標準差不匹配</li>
    <li><span style='color:#8a6d3b;'><strong>資料不足</strong></span>：樣本數少於5個</li>
  </ul>
</div>
        """
        notice_label_zh.setText(notice_html_zh)
        right_layout.addWidget(notice_label_zh)
        
        formula_label_zh = QtWidgets.QLabel()
        formula_label_zh.setWordWrap(True)
        formula_label_zh.setTextFormat(QtCore.Qt.TextFormat.RichText)
        formula_label_zh.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        formula_html_zh = """
<div style="background-color:#e8f4f8; padding:15px; border-radius:6px; font-size:13px; line-height:1.6; font-family:Microsoft JhengHei; border: 2px solid #344CB7;">
  <strong style='font-size:15px;'>📘 計算公式</strong>
  <table style="font-size:12px; margin-top:12px; font-family:Microsoft JhengHei; width:100%;">
    <tr>
      <td style="vertical-align:top; padding:8px; padding-right:12px; width:35%;"><strong>平均值匹配指數：</strong></td>
      <td style="padding:8px;">
        <u>兩組比較：</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>|μ₁ − μ₂| / min(σ₁, σ₂)</code><br><br>
        <u>多組比較：</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>|μ − median(μ)| / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding:8px; padding-right:12px;"><strong>標準差匹配指數：</strong></td>
      <td style="padding:8px;">
        <u>兩組比較：</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>σ / min(σ₁, σ₂)</code><br><br>
        <u>多組比較：</u><br>
        <code style='background:#fff; padding:2px 6px; border-radius:3px;'>σ / median(σ)</code>
      </td>
    </tr>
    <tr>
      <td style="vertical-align:top; padding:8px; padding-right:12px;"><strong>K 值：</strong></td>
      <td style="padding:8px;">
        <code style='background:#fff; padding:6px 10px; border-radius:3px; display:block; line-height:1.8;'>
          n = 樣本數<br>
          n ≤ 4：不進行比較<br>
          5 ≤ n ≤ 10：K = 1.73<br>
          11 ≤ n ≤ 120：K = 1.414<br>
          n > 120：K = 1.15
        </code>
      </td>
    </tr>
  </table>
  <div style="margin-top:15px; padding:12px; background:#fff; border-radius:4px; font-size:12px; color:#344CB7;">
    <strong>【篩選模式計算說明】</strong><br>
    <ul style="margin:8px 0 8px 20px; padding-left:0; line-height:1.8;">
      <li><strong>平均值/標準差/樣本數：</strong>對每個 matching_group，取「指定日期前一個月內」的資料，不足時補滿至指定筆數（預設5筆，可由UI調整），再計算 mean/std/count。</li>
      <li><strong>Median(sigma)：</strong>對每個 matching_group，取「指定日期前六個月內」的資料，不足時補滿至指定筆數（預設5筆），對每組計算 std 後取中位數。</li>
      <li><strong>圖表顯示：</strong>基於「一個月內（補滿至指定筆數）」的資料。</li>
      <li><strong>多筆資料：</strong>若同一組同一時間有多筆資料，皆納入計算。</li>
    </ul>
    <span style="color:#8a6d3b; font-style:italic;">• 全算模式直接使用所有資料進行分組計算，不進行補滿</span>
  </div>
</div>
        """
        formula_label_zh.setText(formula_html_zh)
        right_layout.addWidget(formula_label_zh)
        right_layout.addStretch()
        
        # 加入左右兩側到內容佈局
        content_layout.addWidget(left_widget)
        content_layout.addWidget(right_widget)
        
        scroll.setWidget(content_widget)
        layout.addWidget(scroll)
        
        # 關閉按鈕
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()
        close_btn = QtWidgets.QPushButton(tr("close"))
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #344CB7;
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #577BC1; }
            QPushButton:pressed { background-color: #000957; }
        """)
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)


class ToolMatchingSettingsDialog(QtWidgets.QDialog):
    """Tool Matching 設定對話框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("settings"))
        self.setMinimumWidth(500)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        
        # 分組：閾值設定
        threshold_group = QtWidgets.QGroupBox(tr("threshold_settings"))
        threshold_layout = QtWidgets.QVBoxLayout(threshold_group)
        
        mean_layout = QtWidgets.QHBoxLayout()
        self.mean_index_checkbox = QtWidgets.QCheckBox()
        self.mean_threshold_label = QtWidgets.QLabel(tr("mean_index_threshold"))
        self.mean_index_threshold_spin = QtWidgets.QDoubleSpinBox()
        self.mean_index_threshold_spin.setRange(0, 10)
        self.mean_index_threshold_spin.setValue(1.0)
        self.mean_index_threshold_spin.setSingleStep(0.1)
        self.mean_index_threshold_spin.setEnabled(False)
        mean_layout.addWidget(self.mean_index_checkbox)
        mean_layout.addWidget(self.mean_threshold_label)
        mean_layout.addWidget(self.mean_index_threshold_spin)
        mean_layout.addStretch()
        self.mean_index_checkbox.stateChanged.connect(
            lambda: self.mean_index_threshold_spin.setEnabled(self.mean_index_checkbox.isChecked())
        )
        threshold_layout.addLayout(mean_layout)
        
        sigma_layout = QtWidgets.QHBoxLayout()
        self.sigma_index_checkbox = QtWidgets.QCheckBox()
        self.sigma_threshold_label = QtWidgets.QLabel(tr("sigma_index_threshold"))
        self.sigma_index_threshold_spin = QtWidgets.QDoubleSpinBox()
        self.sigma_index_threshold_spin.setRange(0, 10)
        self.sigma_index_threshold_spin.setValue(2.0)
        self.sigma_index_threshold_spin.setSingleStep(0.1)
        self.sigma_index_threshold_spin.setEnabled(False)
        sigma_layout.addWidget(self.sigma_index_checkbox)
        sigma_layout.addWidget(self.sigma_threshold_label)
        sigma_layout.addWidget(self.sigma_index_threshold_spin)
        sigma_layout.addStretch()
        self.sigma_index_checkbox.stateChanged.connect(
            lambda: self.sigma_index_threshold_spin.setEnabled(self.sigma_index_checkbox.isChecked())
        )
        threshold_layout.addLayout(sigma_layout)
        
        layout.addWidget(threshold_group)
        
        # 分組：數據處理設定
        data_group = QtWidgets.QGroupBox(tr("data_processing_settings"))
        data_layout = QtWidgets.QVBoxLayout(data_group)
        
        fillnum_layout = QtWidgets.QHBoxLayout()
        self.fillnum_checkbox = QtWidgets.QCheckBox()
        self.fillnum_label = QtWidgets.QLabel(tr("fill_sample_size"))
        self.fillnum_spin = QtWidgets.QSpinBox()
        self.fillnum_spin.setMinimum(1)
        self.fillnum_spin.setMaximum(100)
        self.fillnum_spin.setValue(5)
        self.fillnum_spin.setEnabled(False)
        fillnum_layout.addWidget(self.fillnum_checkbox)
        fillnum_layout.addWidget(self.fillnum_label)
        fillnum_layout.addWidget(self.fillnum_spin)
        fillnum_layout.addStretch()
        self.fillnum_checkbox.stateChanged.connect(
            lambda: self.fillnum_spin.setEnabled(self.fillnum_checkbox.isChecked())
        )
        data_layout.addLayout(fillnum_layout)
        
        filter_layout = QtWidgets.QHBoxLayout()
        self.filter_mode_label = QtWidgets.QLabel(tr("data_filter_mode"))
        self.filter_mode_combo = QtWidgets.QComboBox()
        self.filter_mode_combo.addItems([
            tr("all_data"),
            tr("specified_date"),
            tr("latest_entry")
        ])
        self.filter_mode_combo.setFixedWidth(260)
        filter_layout.addWidget(self.filter_mode_label)
        filter_layout.addWidget(self.filter_mode_combo)
        filter_layout.addStretch()
        data_layout.addLayout(filter_layout)
        
        date_layout = QtWidgets.QHBoxLayout()
        self.base_date_label = QtWidgets.QLabel(tr("specified_base_date"))
        self.date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate())
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setEnabled(False)
        date_layout.addWidget(self.base_date_label)
        date_layout.addWidget(self.date_edit)
        date_layout.addStretch()
        self.filter_mode_combo.currentIndexChanged.connect(
            lambda idx: self.date_edit.setEnabled(idx == 1)
        )
        data_layout.addLayout(date_layout)
        
        layout.addWidget(data_group)
        
        # 按鈕
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()
        
        cancel_btn = QtWidgets.QPushButton(tr("cancel"))
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        save_btn = QtWidgets.QPushButton(tr("save"))
        save_btn.setStyleSheet("""
            QPushButton { background-color: #344CB7; color: white; border-radius: 8px; padding: 8px 20px; font-weight: bold; }
            QPushButton:hover { background-color: #577BC1; }
            QPushButton:pressed { background-color: #000957; }
        """)
        save_btn.clicked.connect(self.accept)
        button_layout.addWidget(save_btn)
        
        layout.addLayout(button_layout)
    
    def get_settings(self):
        return {
            'mean_index_enabled': self.mean_index_checkbox.isChecked(),
            'mean_index_threshold': self.mean_index_threshold_spin.value(),
            'sigma_index_enabled': self.sigma_index_checkbox.isChecked(),
            'sigma_index_threshold': self.sigma_index_threshold_spin.value(),
            'fillnum_enabled': self.fillnum_checkbox.isChecked(),
            'fillnum_value': self.fillnum_spin.value(),
            'filter_mode': self.filter_mode_combo.currentIndex(),
            'base_date': self.date_edit.date()
        }
    
    def set_settings(self, settings):
        self.mean_index_checkbox.setChecked(settings.get('mean_index_enabled', False))
        self.mean_index_threshold_spin.setValue(settings.get('mean_index_threshold', 1.0))
        self.sigma_index_checkbox.setChecked(settings.get('sigma_index_enabled', False))
        self.sigma_index_threshold_spin.setValue(settings.get('sigma_index_threshold', 2.0))
        self.fillnum_checkbox.setChecked(settings.get('fillnum_enabled', False))
        self.fillnum_spin.setValue(settings.get('fillnum_value', 5))
        self.filter_mode_combo.setCurrentIndex(settings.get('filter_mode', 0))
        self.date_edit.setDate(settings.get('base_date', QtCore.QDate.currentDate()))


class ToolMatchingWidget(QtWidgets.QWidget):
    """
    Tool Matching Analysis Tool (Auto-OOB Style):
    - Auto read All_Chart_Information.xlsx
    - Auto load matching CSVs from raw_charts/
    - Auto apply Characteristics from master sheet
    - Perform mean/sigma matching checks based on characteristic
    - Display non-matching results and export to Excel
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.translator = get_translator()
        self.translator.register_observer(self)
        
        font = QtGui.QFont("Microsoft JhengHei")
        font.setPointSize(10)
        self.setFont(font)
        
        self.tool_matching_settings = {
            'mean_index_enabled': False,
            'mean_index_threshold': 1.0,
            'sigma_index_enabled': False,
            'sigma_index_threshold': 2.0,
            'fillnum_enabled': False,
            'fillnum_value': 5,
            'filter_mode': 0,
            'base_date': QtCore.QDate.currentDate()
        }
        
        self.current_export_data = None
        self.init_ui()

    def _get_resource_path(self, relative_path):
        import sys, os
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def _find_matching_file(self, directory, group_name, chart_name):
        import os, re
        pattern = re.compile(rf"{re.escape(str(group_name))}_{re.escape(str(chart_name))}(?:_\d+_\d+)?\.csv$")
        if not os.path.exists(directory):
            return None
        for filename in os.listdir(directory):
            if pattern.match(filename):
                return os.path.join(directory, filename)
        return None

    def refresh_ui_texts(self):
        self.setWindowTitle(tr("tool_matching_title"))
        self.title_label.setText(f"<h2 style='color:#34495E;'>{tr('tool_matching_title')}</h2>")
        self.settings_button.setText(f"⚙️ {tr('settings')}")
        self.formula_btn.setText(f"📊 {tr('formula_explanation')}")
        self.run_btn.setText(f"▶ {tr('run_analysis')}")
        self.export_btn.setText(tr('export_to_excel'))
        if self.status_label.text() in ["Ready", "準備就緒"]:
            self.status_label.setText(tr("ready"))
        
        self.result_table.setHorizontalHeaderLabels([
            "View Details", "Abnormal Type", tr("group_name"), tr("chart_name"), tr("matching_group"), 
            tr("mean_index"), tr("sigma_index"), tr("k_value"),
            tr("mean"), tr("sigma"), tr("mean_median"), 
            tr("sigma_median"), tr("sample_size"), "Characteristic"
        ])
    
    def init_ui(self):
        self.setWindowTitle(tr("tool_matching_title"))
        self.resize(1200, 800)
        self.main_layout = QtWidgets.QVBoxLayout(self)
        self.setLayout(self.main_layout)

        # --- Top Control Area ---
        top_layout_widget = QtWidgets.QWidget()
        top_layout = QtWidgets.QVBoxLayout(top_layout_widget)

        self.title_label = QtWidgets.QLabel(f"<h2 style='color:#34495E;'>{tr('tool_matching_title')}</h2>")
        title_font = QtGui.QFont("Microsoft JhengHei")
        title_font.setPointSize(16)
        self.title_label.setFont(title_font)
        top_layout.addWidget(self.title_label)

        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(10)
        
        btn_font = QtGui.QFont("Microsoft JhengHei")
        btn_font.setBold(True)
        btn_font.setPointSize(11)

        self.settings_button = QtWidgets.QPushButton(f"⚙️ {tr('settings')}")
        self.settings_button.setMinimumHeight(45)
        self.settings_button.setMinimumWidth(120)
        self.settings_button.setFont(btn_font)
        self.settings_button.setStyleSheet("""
            QPushButton { background-color: #f8f9fa; color: #333; border: 2px solid #dee2e6; border-radius: 8px; padding: 8px 15px; font-weight: bold; }
            QPushButton:hover { background-color: #e9ecef; border-color: #adb5bd; }
            QPushButton:pressed { background-color: #dee2e6; }
        """)
        self.settings_button.clicked.connect(self.open_tool_matching_settings)
        button_layout.addWidget(self.settings_button)
        
        self.formula_btn = QtWidgets.QPushButton(f"📊 {tr('formula_explanation')}")
        self.formula_btn.setMinimumHeight(45)
        self.formula_btn.setMinimumWidth(120)
        self.formula_btn.setFont(btn_font)
        self.formula_btn.setStyleSheet(self.settings_button.styleSheet())
        self.formula_btn.clicked.connect(self.open_formula_explanation)
        button_layout.addWidget(self.formula_btn)
        
        self.run_btn = QtWidgets.QPushButton(f"▶ {tr('run_analysis')}")
        self.run_btn.setMinimumHeight(45)
        self.run_btn.setFont(btn_font)
        self.run_btn.setStyleSheet("""
            QPushButton { background-color: #344CB7; color: white; border: none; border-radius: 8px; padding: 10px 20px; font-weight: bold; }
            QPushButton:hover { background-color: #577BC1; }
            QPushButton:pressed { background-color: #000957; }
            QPushButton:disabled { background-color: #cccccc; color: #666666; }
        """)
        self.run_btn.clicked.connect(self.run_analysis)
        button_layout.addWidget(self.run_btn)

        self.export_btn = QtWidgets.QPushButton(tr('export_to_excel'))
        self.export_btn.setMinimumHeight(45)
        self.export_btn.setFont(btn_font)
        self.export_btn.setStyleSheet("""
            QPushButton { background-color: #217346; color: white; border: none; border-radius: 8px; padding: 10px 20px; font-weight: bold; }
            QPushButton:hover { background-color: #2e9e5e; }
            QPushButton:pressed { background-color: #145c30; }
            QPushButton:disabled { background-color: #cccccc; color: #666666; }
        """)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_data)
        button_layout.addWidget(self.export_btn)

        button_layout.addStretch()
        top_layout.addLayout(button_layout)

        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setFixedHeight(20)
        self.progress_bar.setVisible(False)
        top_layout.addWidget(self.progress_bar)

        self.status_label = QtWidgets.QLabel(tr("ready"))
        self.status_label.setFont(QtGui.QFont("Microsoft JhengHei", 10))
        top_layout.addWidget(self.status_label)

        self.main_layout.addWidget(top_layout_widget)

        # --- 結果表格 ---
        self.result_table = QtWidgets.QTableWidget()
        self.result_table.setColumnCount(14) 
        self.result_table.setHorizontalHeaderLabels([
            "View Details", "Abnormal Type", tr("group_name"), tr("chart_name"), tr("matching_group"), 
            tr("mean_index"), tr("sigma_index"), tr("k_value"),
            tr("mean"), tr("sigma"), tr("mean_median"), 
            tr("sigma_median"), tr("sample_size"), "Characteristic"
        ])
        self.result_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.result_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.setStyleSheet("""
            QTableWidget { gridline-color: #d0d0d0; }
            QHeaderView::section { background-color: #344CB7; color: white; padding: 4px; font-weight: bold; }
            QTableWidget::item { background: transparent; }
            QTableWidget::item:selected { background: #e6f0fa !important; color: #222 !important; }
        """)
        self.main_layout.addWidget(self.result_table, 1)

    def open_tool_matching_settings(self):
        dialog = ToolMatchingSettingsDialog(self)
        dialog.set_settings(self.tool_matching_settings)
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self.tool_matching_settings = dialog.get_settings()
    
    def open_formula_explanation(self):
        dialog = FormulaExplanationDialog(self)
        dialog.exec()

    def get_k_value(self, n):
        if n <= 4:
            return "No Comparison"
        elif 5 <= n <= 10:
            return 1.73
        elif 11 <= n <= 120:
            return 1.414
        else:
            return 1.15

    def calculate_mean_index(self, mean1, mean2, min_sigma, characteristic):
        if min_sigma <= 0:
            return float('inf')
        if characteristic == 'Bigger':
            return (mean2 - mean1) / min_sigma
        elif characteristic in ['Smaller', 'Sigma']:
            return (mean1 - mean2) / min_sigma
        else:
            return abs(mean1 - mean2) / min_sigma

    def run_analysis(self):
        info_path = self._get_resource_path('input/All_Chart_Information.xlsx')
        raw_dir = self._get_resource_path('input/raw_charts/')

        if not os.path.exists(info_path) or not os.path.exists(raw_dir):
            self.status_label.setText("找不到輸入檔案：請確保 input/All_Chart_Information.xlsx 與 input/raw_charts/ 存在。")
            QtWidgets.QMessageBox.critical(self, "Error", "無法找到所需的輸入檔案。")
            return

        try:
            all_charts_info = pd.read_excel(info_path, sheet_name='Chart', engine='openpyxl')
        except Exception as e:
            self.status_label.setText(f"讀取 Excel 失敗: {e}")
            return

        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting file processing...")
        QtWidgets.QApplication.processEvents()

        filter_mode = self.tool_matching_settings.get('filter_mode', 0)
        base_date = self.tool_matching_settings.get('base_date', QtCore.QDate.currentDate()).toPyDate() if filter_mode == 1 else None
        fill_num = self.tool_matching_settings.get('fillnum_value', 5)

        results = []
        self.chart_figures = {}
        total_charts = len(all_charts_info)

        for i, row in all_charts_info.iterrows():
            gname = str(row.get('GroupName', 'Unknown'))
            cname = str(row.get('ChartName', 'Unknown'))
            
            # 🔥 從 All_Chart_Information.xlsx 讀取屬性，預設為 Nominal
            characteristic = str(row.get('Characteristics', 'Nominal'))

            self.progress_bar.setValue(int(((i + 1) / total_charts) * 100))
            self.status_label.setText(f"Processing {i+1}/{total_charts}: {gname}_{cname}")
            QtWidgets.QApplication.processEvents()

            filepath = self._find_matching_file(raw_dir, gname, cname)
            if not filepath or not os.path.exists(filepath):
                continue

            try:
                subdf = pd.read_csv(filepath)
            except Exception:
                continue

            if subdf.empty or 'point_val' not in subdf.columns or 'point_time' not in subdf.columns:
                continue

            # 🔥 自動處理舊版的 ByTool 與新版的 Matching 欄位
            if 'Matching' in subdf.columns:
                subdf['matching_group'] = subdf['Matching']
            elif 'ByTool' in subdf.columns:
                subdf['matching_group'] = subdf['ByTool']
            else:
                subdf['matching_group'] = 'Unknown'

            subdf['characteristic'] = characteristic
            subdf['GroupName'] = gname
            subdf['ChartName'] = cname
            subdf["point_time"] = pd.to_datetime(subdf["point_time"], errors='coerce')
            subdf = subdf.dropna(subset=['point_time', 'point_val'])

            if subdf.empty:
                continue

            if filter_mode == 0:
                group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count']).reset_index()
                n_groups = len(group_stats)
                if n_groups == 2:
                    self._analyze_two_groups(group_stats, gname, cname, characteristic, results)
                else:
                    self._analyze_multiple_groups(subdf, group_stats, gname, cname, characteristic, results)
                self._create_boxplot_single(subdf, gname, cname)

            elif filter_mode == 1 or filter_mode == 2:
                if filter_mode == 1:
                    mean_end = pd.Timestamp(base_date)
                    sigma_end = pd.Timestamp(base_date)
                else:
                    latest_time = subdf["point_time"].max()
                    mean_end = latest_time
                    sigma_end = latest_time

                mean_start = mean_end - pd.DateOffset(months=1)
                sigma_start = sigma_end - pd.DateOffset(months=6)

                mean_df = subdf[(subdf["point_time"] > mean_start) & (subdf["point_time"] <= mean_end)].copy()
                sigma_df = subdf[(subdf["point_time"] > sigma_start) & (subdf["point_time"] <= sigma_end)].copy()
                min_time = subdf["point_time"].min()

                # 補齊 mean_df
                for mg in subdf["matching_group"].unique():
                    mg_mean = mean_df[mean_df["matching_group"] == mg]
                    if len(mg_mean) < fill_num:
                        all_mg = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                        cur_start = mean_start
                        while len(mg_mean) < fill_num and cur_start > min_time:
                            cur_start = cur_start - pd.Timedelta(days=7)
                            mg_mean = all_mg[(all_mg["point_time"] > cur_start) & (all_mg["point_time"] <= mean_end)]
                        mean_df = pd.concat([mean_df, mg_mean]).drop_duplicates()

                # 補齊 sigma_df
                for mg in subdf["matching_group"].unique():
                    mg_sigma = sigma_df[sigma_df["matching_group"] == mg]
                    if len(mg_sigma) < fill_num:
                        all_mg = subdf[subdf["matching_group"] == mg].sort_values("point_time")
                        cur_start = sigma_start
                        while len(mg_sigma) < fill_num and cur_start > min_time:
                            cur_start = cur_start - pd.Timedelta(days=14)
                            mg_sigma = all_mg[(all_mg["point_time"] > cur_start) & (all_mg["point_time"] <= sigma_end)]
                        sigma_df = pd.concat([sigma_df, mg_sigma]).drop_duplicates()

                mean_stats = mean_df.groupby("matching_group")["point_val"].agg(['mean', 'count']).reset_index()
                sigma_stats = sigma_df.groupby("matching_group")["point_val"].agg(['std']).reset_index()
                group_stats = pd.merge(mean_stats, sigma_stats, on="matching_group", how="outer")
                group_stats = group_stats.fillna({"mean": 0, "std": 0, "count": 0})

                n_groups = len(group_stats)
                if n_groups == 2:
                    self._analyze_two_groups(group_stats, gname, cname, characteristic, results)
                else:
                    self._analyze_multiple_groups_time(mean_df, sigma_df, group_stats, gname, cname, characteristic, results)
                
                self._create_boxplot_single(mean_df if not mean_df.empty else subdf, gname, cname)

        self.progress_bar.setVisible(False)
        self._display_results(results)

    def _analyze_two_groups(self, group_stats, gname, cname, characteristic, results):
        row1 = group_stats.iloc[0]
        row2 = group_stats.iloc[1]

        group1 = row1["matching_group"]
        group2 = row2["matching_group"]
        mean1, std1, n1 = row1["mean"], row1["std"], row1["count"]
        mean2, std2, n2 = row2["mean"], row2["std"], row2["count"]

        min_sigma = min(std1, std2)

        if n1 < 5 or n2 < 5:
            results.append([gname, cname, group1, 'group_all', 'Insufficient Data', 'Insufficient Data', self.get_k_value(n1), mean1, std1, mean2, min_sigma, n1, characteristic])
            results.append([gname, cname, group2, 'group_all', 'Insufficient Data', 'Insufficient Data', self.get_k_value(n2), mean2, std2, mean1, min_sigma, n2, characteristic])
            return

        k1, k2 = self.get_k_value(n1), self.get_k_value(n2)

        if min_sigma > 0:
            mean_index_1 = self.calculate_mean_index(mean1, mean2, min_sigma, characteristic)
            sigma_index_1 = std1 / min_sigma
        else:
            if len(set([round(m, 8) for m in [mean1, mean2]])) == 1:
                mean_index_1, sigma_index_1 = 0, 0
            else:
                mean_index_1, sigma_index_1 = float('inf'), float('inf')

        results.append([gname, cname, group1, 'group_all', 
            'Insufficient Data' if k1=="No Comparison" else round(mean_index_1, 2), 
            'Insufficient Data' if k1=="No Comparison" else round(sigma_index_1, 2),
            k1 if isinstance(k1, str) else round(k1, 2), 
            round(mean1, 2), round(std1, 2), round(mean2, 2), round(min_sigma, 2), n1, characteristic])

        if min_sigma > 0:
            mean_index_2 = self.calculate_mean_index(mean2, mean1, min_sigma, characteristic)
            sigma_index_2 = std2 / min_sigma
        else:
            if len(set([round(m, 8) for m in [mean1, mean2]])) == 1:
                mean_index_2, sigma_index_2 = 0, 0
            else:
                mean_index_2, sigma_index_2 = float('inf'), float('inf')

        results.append([gname, cname, group2, 'group_all', 
            'Insufficient Data' if k2=="No Comparison" else round(mean_index_2, 2), 
            'Insufficient Data' if k2=="No Comparison" else round(sigma_index_2, 2),
            k2 if isinstance(k2, str) else round(k2, 2), 
            round(mean2, 2), round(std2, 2), round(mean1, 2), round(min_sigma, 2), n2, characteristic])

    def _analyze_multiple_groups(self, subdf, group_stats, gname, cname, characteristic, results):
        valid_stats = group_stats[group_stats['count'] >= 5]
        
        if valid_stats.shape[0] == 2:
            self._analyze_two_groups(valid_stats, gname, cname, characteristic, results)
            for i, row in group_stats[group_stats['count'] < 5].iterrows():
                results.append([gname, cname, row["matching_group"], "group_all", 'Insufficient Data', 'Insufficient Data', self.get_k_value(row["count"]), row["mean"], row["std"], '-', '-', row["count"], characteristic])
            return
        
        if valid_stats.shape[0] <= 1:
            for i, row in group_stats.iterrows():
                results.append([gname, cname, row["matching_group"], "group_all", 'Insufficient Data', 'Insufficient Data', self.get_k_value(row["count"]), row["mean"], row["std"], '-', '-', row["count"], characteristic])
            return

        mean_median = valid_stats['mean'].median() if not valid_stats.empty else 0
        median_sigma = valid_stats['std'].median() if not valid_stats.empty else 0

        for i, row in group_stats.iterrows():
            group, mean, std, n = row["matching_group"], row["mean"], row["std"], row["count"]
            if n < 5:
                results.append([gname, cname, group, "group_all", 'Insufficient Data', 'Insufficient Data', self.get_k_value(n), mean, std, mean_median, median_sigma, n, characteristic])
                continue

            if median_sigma > 0:
                mean_index = self.calculate_mean_index(mean, mean_median, median_sigma, characteristic)
                sigma_index = std / median_sigma
            else:
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    mean_index, sigma_index = 0, 0
                else:
                    mean_index, sigma_index = float('inf'), float('inf')

            K = self.get_k_value(n)
            results.append([gname, cname, group, "group_all", 
                'Insufficient Data' if K=="No Comparison" else round(mean_index, 2), 
                'Insufficient Data' if K=="No Comparison" else round(sigma_index, 2), 
                K if isinstance(K, str) else round(K, 2), 
                round(mean, 2), round(std, 2), round(mean_median, 2), round(median_sigma, 2), n, characteristic])

    def _analyze_multiple_groups_time(self, mean_df, sigma_df, group_stats, gname, cname, characteristic, results):
        valid_mean_df = mean_df.groupby("matching_group").filter(lambda x: len(x) >= 5)
        sigma_by_group = sigma_df.groupby("matching_group")["point_val"].std()
        valid_groups = group_stats[group_stats['count'] >= 5]['matching_group']
        valid_sigma = sigma_by_group[valid_groups] if not valid_groups.empty else pd.Series(dtype=float)
        
        if len(valid_groups) == 2:
            valid_stats = group_stats[group_stats['count'] >= 5]
            self._analyze_two_groups(valid_stats, gname, cname, characteristic, results)
            for i, row in group_stats[group_stats['count'] < 5].iterrows():
                results.append([gname, cname, row["matching_group"], "group_all", 'Insufficient Data', 'Insufficient Data', self.get_k_value(row["count"]), row["mean"], row["std"], '-', '-', row["count"], characteristic])
            return
        
        if len(valid_groups) <= 1:
            for i, row in group_stats.iterrows():
                results.append([gname, cname, row["matching_group"], "group_all", 'Insufficient Data', 'Insufficient Data', self.get_k_value(row["count"]), row["mean"], row["std"], '-', '-', row["count"], characteristic])
            return
            
        mean_median = valid_mean_df["point_val"].median() if not valid_mean_df.empty else 0
        median_sigma = valid_sigma.median() if not valid_sigma.empty else 0
        
        for i, row in group_stats.iterrows():
            group, mean, std, n = row["matching_group"], row["mean"], row["std"], row["count"]
            if n < 5:
                results.append([gname, cname, group, "group_all", 'Insufficient Data', 'Insufficient Data', self.get_k_value(n), mean, std, mean_median, median_sigma, n, characteristic])
                continue
                
            if median_sigma > 0:
                mean_index = self.calculate_mean_index(mean, mean_median, median_sigma, characteristic)
                sigma_index = std / median_sigma
            else:
                all_means = group_stats['mean'].tolist() if not group_stats.empty else [mean]
                if len(set([round(m, 8) for m in all_means])) == 1:
                    mean_index, sigma_index = 0, 0
                else:
                    mean_index, sigma_index = float('inf'), float('inf')
                    
            K = self.get_k_value(n)
            results.append([gname, cname, group, "group_all", 
                'Insufficient Data' if K=="No Comparison" else round(mean_index, 2), 
                'Insufficient Data' if K=="No Comparison" else round(sigma_index, 2), 
                K if isinstance(K, str) else round(K, 2), 
                round(mean, 2), round(std, 2), round(mean_median, 2), round(median_sigma, 2), n, characteristic])

    def _create_boxplot_single(self, subdf, gname, cname):
        try:
            import matplotlib.pyplot as plt
            from matplotlib import cm
            import numpy as np
        except ImportError:
            return

        unique_groups = sorted(subdf["matching_group"].unique(), key=lambda x: str(x))
        labels = [str(mg) for mg in unique_groups]

        if subdf.empty or not any(len(grp["point_val"]) > 0 for _, grp in subdf.groupby("matching_group")):
            self.chart_figures[(gname, cname)] = {'scatter': None, 'box': None}
            return

        box_data = [subdf[subdf["matching_group"] == mg]["point_val"].values for mg in unique_groups]
        group_stats = subdf.groupby("matching_group")["point_val"].agg(['mean', 'std', 'count'])
        colors = cm.tab10(np.linspace(0, 1, len(unique_groups)))

        # 1. SPC 風格圖表
        scatter_fig, scatter_ax = plt.subplots(figsize=(7, 4.5))
        x_position = 0
        for i, mg in enumerate(unique_groups):
            group_data = subdf[subdf["matching_group"] == mg].sort_values("point_time")
            if not group_data.empty:
                x_vals = np.arange(x_position, x_position + len(group_data))
                y_vals = group_data["point_val"].values
                scatter_ax.scatter(x_vals, y_vals, color=colors[i], alpha=0.8, s=40, label=f'{mg}', zorder=3)
                scatter_ax.plot(x_vals, y_vals, color=colors[i], alpha=0.5, linewidth=1, zorder=2)
                if i < len(unique_groups) - 1:
                    scatter_ax.axvline(x=x_position + len(group_data) - 0.5, color='gray', linestyle='-', alpha=0.3, zorder=1)
                x_position += len(group_data)

        scatter_ax.set_title(f"SPC Chart: {gname} - {cname}", fontsize=10)
        scatter_ax.set_xlabel("Sample Sequence (Grouped by Matching Group)")
        scatter_ax.set_ylabel("Point Value")
        scatter_ax.grid(True, linestyle='--', alpha=0.3, zorder=0)

        if unique_groups:
            group_positions, x_pos = [], 0
            for mg in unique_groups:
                group_size = len(subdf[subdf["matching_group"] == mg])
                group_positions.append(x_pos + group_size/2 - 0.5)
                x_pos += group_size
            scatter_ax.set_xticks(group_positions)
            scatter_ax.set_xticklabels(labels, rotation=0, ha='center')
            scatter_ax.tick_params(axis='x', which='minor', bottom=True, top=False)

        scatter_ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')
        scatter_fig.tight_layout()

        # 2. 盒鬚圖
        box_fig, box_ax = plt.subplots(figsize=(7, 4.5))
        if box_data:
            bp = box_ax.boxplot(box_data, labels=labels, patch_artist=True, widths=0.6)
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)
            legend_labels = [
                f"{label}: μ={group_stats.loc[mg, 'mean']:.2f}, σ={group_stats.loc[mg, 'std']:.2f}, n={int(group_stats.loc[mg, 'count'])}"
                for label, mg in zip(labels, unique_groups)
            ]
            box_ax.legend([bp["boxes"][i] for i in range(len(labels))], legend_labels, loc='upper left', bbox_to_anchor=(1.02, 1), fontsize='small')

        box_ax.set_title(f"Boxplot: {gname} - {cname}", fontsize=10)
        box_ax.set_xlabel("Matching Group")
        box_ax.set_ylabel("Point Value")
        box_ax.grid(True, linestyle='--', alpha=0.6)
        box_fig.subplots_adjust(right=0.7)
        box_fig.tight_layout()

        self.chart_figures[(gname, cname)] = {'scatter': scatter_fig, 'box': box_fig}
        plt.close(scatter_fig)
        plt.close(box_fig)

    def _display_results(self, results):
            self.report_data = {}
            all_table_rows = []
            abnormal_ui_rows = []

            for row in results:
                gname, cname = row[0], row[1]
                key = f"{gname}_{cname}"
                
                if key not in self.report_data:
                    self.report_data[key] = {"GroupName": gname, "ChartName": cname, "groups": {}}
                
                group1, group2 = row[2], row[3]
                mean_index, sigma_index = row[4], row[5]
                k_value, mean, sigma, mean_median, sigma_median, n, characteristic = row[6:13] if len(row) >= 13 else [""] * 6 + [row[6] if len(row) > 6 else ""]
                
                stats_dict = {
                    "mean_matching_index": mean_index, "sigma_matching_index": sigma_index, "K": k_value,
                    "mean": mean, "sigma": sigma, "mean_median": mean_median, "sigma_median": sigma_median,
                    "samplesize": n, "characteristic": characteristic
                }
                if group2 == "group_all":
                    self.report_data[key]["groups"][group1] = stats_dict
                else:
                    if group1 not in self.report_data[key]["groups"]: self.report_data[key]["groups"][group1] = {}
                    self.report_data[key]["groups"][group1][group2] = stats_dict

            for key, data in self.report_data.items():
                gname, cname = data["GroupName"], data["ChartName"]
                for group_id, stats in data["groups"].items():
                    m_idx, s_idx, k_val = stats.get("mean_matching_index", ""), stats.get("sigma_matching_index", ""), stats.get("K", "")
                    
                    is_abnormal = False
                    abnormal_type = ""
                    is_data_insufficient = m_idx == 'Insufficient Data' or s_idx == 'Insufficient Data' or k_val == 'No Comparison'
                    
                    if not is_data_insufficient:
                        try:
                            mean_threshold = self.tool_matching_settings.get('mean_index_threshold', 1.0) if self.tool_matching_settings.get('mean_index_enabled', False) else 1.0
                            sigma_threshold = self.tool_matching_settings.get('sigma_index_threshold', 2.0) if self.tool_matching_settings.get('sigma_index_enabled', False) else (float(k_val) if k_val not in [None, '', 'No Comparison'] else 2.0)
                            mean_abn = float(m_idx) >= mean_threshold
                            sigma_abn = float(s_idx) >= sigma_threshold
                            if mean_abn or sigma_abn:
                                is_abnormal = True
                                if mean_abn and sigma_abn: abnormal_type = "Mean, Sigma"
                                elif mean_abn: abnormal_type = "Mean"
                                elif sigma_abn: abnormal_type = "Sigma"
                        except (ValueError, TypeError):
                            pass
                    
                    n_val = stats.get("samplesize", "")
                    try: n_val = int(float(n_val)) if n_val not in ['', None] else n_val
                    except: pass
                    
                    row_data = [
                        gname, cname, group_id, m_idx, s_idx, k_val,
                        stats.get("mean", ""), stats.get("sigma", ""), stats.get("mean_median", ""), 
                        stats.get("sigma_median", ""), n_val, stats.get("characteristic", "")
                    ]
                    
                    all_table_rows.append([is_abnormal, abnormal_type] + row_data)
                    
                    if is_abnormal:
                        abnormal_ui_rows.append({"key": (gname, cname), "group_id": group_id, "data": [abnormal_type] + row_data})

            self.result_table.setRowCount(len(abnormal_ui_rows))
            for i, item_info in enumerate(abnormal_ui_rows):
                view_button = QtWidgets.QPushButton()
                try: view_button.setIcon(self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_FileDialogContentsView))
                except: view_button.setIcon(self.style().standardIcon(QtWidgets.QStyle.StandardPixmap.SP_DesktopIcon))
                view_button.setToolTip("View Details")
                view_button.setFixedSize(36, 36)
                view_button.setIconSize(QtCore.QSize(22, 22))
                view_button.setStyleSheet("QPushButton { border: none; background: transparent; } QPushButton:hover { background: #e0e7ef; }")
                
                cell_widget = QtWidgets.QWidget()
                layout = QtWidgets.QHBoxLayout(cell_widget)
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                layout.addWidget(view_button)
                view_button.clicked.connect(lambda checked, key=item_info["key"], gid=item_info["group_id"]: self._show_details_dialog(key, gid))
                self.result_table.setCellWidget(i, 0, cell_widget)

                for j, val in enumerate(item_info["data"]):
                    if j in [4,5,6,7,8,9,10]:
                        try:
                            if val not in ['Insufficient Data', 'No Comparison', '', None]: val = f"{float(val):.2f}"
                        except: pass
                    item = QtWidgets.QTableWidgetItem(str(val))
                    item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                    
                    is_abn = False
                    try:
                        if (j == 4 and float(item_info["data"][4]) >= 1) or (j == 5 and float(item_info["data"][5]) >= float(item_info["data"][6])):
                            is_abn = True
                    except: pass
                    if is_abn: item.setForeground(QtGui.QColor("#D32F2F"))
                    self.result_table.setItem(i, j + 1, item)

            self.result_table.resizeColumnsToContents()
            self.result_table.horizontalHeader().setStretchLastSection(True)

            self.current_export_data = all_table_rows
            if all_table_rows:
                self.export_btn.setEnabled(True)
                if len(abnormal_ui_rows) > 0:
                    self.status_label.setText(f"Analysis completed. Found {len(abnormal_ui_rows)} items requiring attention. Click Export to save results.")
                else:
                    self.status_label.setText(f"Analysis completed. All matching groups are Normal. Click Export to save results.")
            else:
                self.status_label.setText("Analysis completed, but no data available.")

    def export_data(self):
        if not self.current_export_data:
            QtWidgets.QMessageBox.warning(self, tr('export_failed'), "No analysis data to export. Please run analysis first.")
            return

        total = len(self.current_export_data)

        self.export_dialog = QtWidgets.QProgressDialog(tr('exporting_charts'), tr('cancel'), 0, total, self)
        self.export_dialog.setWindowTitle(tr('export_progress'))
        self.export_dialog.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
        self.export_dialog.setMinimumWidth(420)
        self.export_dialog.setAutoReset(False)
        self.export_dialog.setAutoClose(False)
        self.export_dialog.setMinimumDuration(0)
        self.export_dialog.setValue(0)
        self.export_dialog.show()
        QtWidgets.QApplication.processEvents()

        QtCore.QTimer.singleShot(100, self._process_export)

    def _process_export(self):
        cancelled = [False]

        def on_cancel():
            cancelled[0] = True
        self.export_dialog.canceled.connect(on_cancel)

        def progress_callback(current, label_text=None):
            if cancelled[0]:
                return False
            self.export_dialog.setValue(current)
            if label_text:
                self.export_dialog.setLabelText(label_text)
            QtWidgets.QApplication.processEvents()
            return not cancelled[0]

        try:
            output_path = self._export_to_excel(self.current_export_data, progress_callback=progress_callback)
        except Exception as e:
            self.export_dialog.canceled.disconnect(on_cancel)
            self.export_dialog.close()
            QtWidgets.QMessageBox.critical(self, tr('export_failed'), f"{tr('export_failed_msg')}\n{e}")
            return

        # 記錄取消狀態後立即 disconnect，避免 close() 觸發 canceled 訊號污染結果
        was_cancelled = cancelled[0]
        self.export_dialog.canceled.disconnect(on_cancel)
        self.export_dialog.close()

        if was_cancelled:
            QtWidgets.QMessageBox.information(self, tr('export_cancelled'), tr('export_cancelled_msg'))
            return

        if output_path:
            QtWidgets.QMessageBox.information(self, tr('export_successful'), f"{tr('export_successful_msg')}\n{output_path}")
            self.status_label.setText(f"{tr('export_successful_msg')} {os.path.basename(output_path)}")
        else:
            QtWidgets.QMessageBox.warning(self, tr('export_failed'), tr('export_failed_msg'))

    def _show_details_dialog(self, chart_key, group_id):
        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
        except ImportError:
            QtWidgets.QMessageBox.warning(self, "Missing Package", "Matplotlib is required to display charts.")
            return

        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"Detailed Information: {chart_key[0]} - {chart_key[1]} | Group: {group_id}")
        dialog.setMinimumSize(1400, 450)

        main_layout = QtWidgets.QVBoxLayout(dialog)
        main_layout.setSpacing(10)

        try:
            stats = self.report_data[f"{chart_key[0]}_{chart_key[1]}"]["groups"][group_id]
        except KeyError:
            QtWidgets.QMessageBox.critical(self, "Error", "Cannot find detailed data for this item.")
            return

        info_group = QtWidgets.QGroupBox("Analysis Data")
        info_v_layout = QtWidgets.QVBoxLayout(info_group)

        m_idx, s_idx, k_val = stats.get("mean_matching_index", ""), stats.get("sigma_matching_index", ""), stats.get("K", "")
        abnormal_type = ""
        if m_idx != 'Insufficient Data' and s_idx != 'Insufficient Data' and k_val != 'No Comparison':
            try:
                m_th = self.tool_matching_settings.get('mean_index_threshold', 1.0) if self.tool_matching_settings.get('mean_index_enabled', False) else 1.0
                s_th = self.tool_matching_settings.get('sigma_index_threshold', 2.0) if self.tool_matching_settings.get('sigma_index_enabled', False) else float(k_val) if k_val not in [None, ''] else 2.0
                m_abn = float(m_idx) >= m_th
                s_abn = float(s_idx) >= s_th
                if m_abn and s_abn: abnormal_type = "Mean, Sigma"
                elif m_abn: abnormal_type = "Mean"
                elif s_abn: abnormal_type = "Sigma"
            except: pass

        gname, cname = chart_key
        n_val = stats.get("samplesize", "")
        try: n_val = int(float(n_val)) if n_val not in ['', None] else n_val
        except: pass
        
        row_values = [
            abnormal_type, gname, cname, group_id, m_idx, s_idx, k_val,
            stats.get("mean", ""), stats.get("sigma", ""), stats.get("mean_median", ""), 
            stats.get("sigma_median", ""), n_val, stats.get("characteristic", "")
        ]

        info_table = QtWidgets.QTableWidget()
        info_table.setColumnCount(13)
        info_table.setHorizontalHeaderLabels([
            "Abnormal Type", "Group Name", "Chart Name", "Matching Group", "Mean Index", "Sigma Index",
            "K", "Mean", "Sigma", "Mean Median", "Sigma Median", "Sample Size", "Characteristic"
        ])
        info_table.setRowCount(1)
        info_table.verticalHeader().setVisible(False)
        info_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)

        for j, value in enumerate(row_values):
            if j in [4,5,6,7,8,9,10]:
                try:
                    if value not in ['Insufficient Data', 'No Comparison', '', None]: value = f"{float(value):.2f}"
                except: pass
            item = QtWidgets.QTableWidgetItem(str(value))
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            # 紅色標記：mean_index(j=4) 或 sigma_index(j=5) 異常時
            try:
                if j == 4 and float(row_values[4]) >= m_th:
                    item.setForeground(QtGui.QColor("#D32F2F"))
                elif j == 5 and float(row_values[5]) >= s_th:
                    item.setForeground(QtGui.QColor("#D32F2F"))
            except (ValueError, TypeError):
                pass
            info_table.setItem(0, j, item)

        info_table.resizeColumnsToContents()
        info_table.setFixedHeight(info_table.horizontalHeader().height() + info_table.rowHeight(0) + 5)
        info_v_layout.addWidget(info_table)
        main_layout.addWidget(info_group)

        charts_container_widget = QtWidgets.QWidget()
        charts_layout = QtWidgets.QHBoxLayout(charts_container_widget)

        if hasattr(self, 'chart_figures') and chart_key in self.chart_figures:
            figures = self.chart_figures[chart_key]
            if figures['scatter'] and figures['box']:
                scatter_fig_copy = pickle.loads(pickle.dumps(figures['scatter']))
                box_fig_copy = pickle.loads(pickle.dumps(figures['box']))
                scatter_canvas = FigureCanvas(scatter_fig_copy)
                box_canvas = FigureCanvas(box_fig_copy)
                scatter_canvas.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Expanding)
                box_canvas.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Expanding)
                charts_layout.addWidget(scatter_canvas)
                charts_layout.addWidget(box_canvas)
            else:
                charts_layout.addWidget(QtWidgets.QLabel("Charts for this item were not generated due to insufficient data."))
        else:
            charts_layout.addWidget(QtWidgets.QLabel("Cannot find corresponding charts."))
        
        main_layout.addWidget(charts_container_widget)
        main_layout.setStretchFactor(info_group, 0)
        main_layout.setStretchFactor(charts_container_widget, 1)

        button_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(dialog.reject)
        main_layout.addWidget(button_box)

        dialog.exec()

    def _export_to_excel(self, all_results, progress_callback=None):
            try:
                import io
                from PIL import Image
                import matplotlib.cm as cm
                from openpyxl.drawing.image import Image as XLImage
                import tempfile
                import shutil
            except ImportError as e:
                print(f"[WARNING] Missing packages required for Excel export: {e}")
                return None

            columns = [
                "Need_matching", "AbnormalType", "GroupName", "ChartName", "matching_group", "mean_matching_index", 
                "sigma_matching_index", "K", "mean", "sigma", "mean_median", "sigma_median", "samplesize", "characteristic"
            ]
            df = pd.DataFrame(all_results, columns=columns)

            output_path = self._get_resource_path('Tool_Matching_Results.xlsx')

            temp_dir = tempfile.mkdtemp()
            
            df.insert(0, "SPC_Chart", "")
            df.insert(1, "BoxPlot", "")

            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            df.to_excel(writer, sheet_name='Tool Matching Results', index=False)
            worksheet = writer.sheets['Tool Matching Results']

            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="344CB7", end_color="344CB7", fill_type="solid")
            header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            worksheet.column_dimensions['A'].width = 70
            worksheet.column_dimensions['B'].width = 70
            abnormal_fill = openpyxl.styles.PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            img_display_width, img_display_height = 450, 250

            has_chart_figures = hasattr(self, 'chart_figures') and self.chart_figures
            
            # 🚀 加速優化 2：圖片快取字典，避免重複存檔 (Savefig) 拖慢速度
            generated_img_cache = {}

            total_rows = len(df)
            for row_idx, row in enumerate(df.iterrows(), start=2):
                _, row_data = row
                if progress_callback is not None:
                    label = f"{row_data.get('GroupName', '')} - {row_data.get('ChartName', '')}"
                    if not progress_callback(row_idx - 1, label):
                        try:
                            writer.close()
                        except Exception:
                            pass
                        return None
                if row_data["Need_matching"]:
                    for cell in worksheet[row_idx]: cell.fill = abnormal_fill

                try:
                    group_name, chart_name = str(row_data["GroupName"]), str(row_data["ChartName"])
                    chart_key = (group_name, chart_name)
                    
                    if has_chart_figures and chart_key in self.chart_figures:
                        chart_data = self.chart_figures[chart_key]
                        
                        # 只有遇到「沒存過圖的 Chart」才執行耗時的 savefig
                        if chart_key not in generated_img_cache:
                            temp_scatter = os.path.join(temp_dir, f"spc_{group_name}_{chart_name}.png")
                            temp_box = os.path.join(temp_dir, f"box_{group_name}_{chart_name}.png")
                            
                            if chart_data.get('scatter'):
                                # 稍微降低 dpi (100 -> 80) 可進一步提升效能，對 Excel 縮圖不影響解析度
                                chart_data['scatter'].savefig(temp_scatter, format='png', bbox_inches='tight', transparent=True, dpi=80)
                            if chart_data.get('box'):
                                chart_data['box'].savefig(temp_box, format='png', bbox_inches='tight', transparent=True, dpi=80)
                                
                            # 紀錄快取路徑
                            generated_img_cache[chart_key] = (temp_scatter, temp_box)
                        
                        # 取出已經存好的圖片路徑，直接貼入 Excel
                        spc_path, box_path = generated_img_cache[chart_key]

                        try:
                            if os.path.exists(spc_path):
                                img1 = XLImage(spc_path)
                                img1.width, img1.height = img_display_width, img_display_height
                                worksheet.add_image(img1, f"A{row_idx}")
                        except Exception: worksheet.cell(row=row_idx, column=1).value = "Error"
                        
                        try:
                            if os.path.exists(box_path):
                                img2 = XLImage(box_path)
                                img2.width, img2.height = img_display_width, img_display_height
                                worksheet.add_image(img2, f"B{row_idx}")
                        except Exception: worksheet.cell(row=row_idx, column=2).value = "Error"
                    else:
                        worksheet.cell(row=row_idx, column=1).value = "No Data"
                        worksheet.cell(row=row_idx, column=2).value = "No Data"
                except Exception: pass

            for i in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[i].height = 190
                
            for col_idx, column in enumerate(worksheet.columns, start=1):
                if col_idx <= 2: continue
                max_len = max([len(str(cell.value)) for cell in column if cell.value] + [0])
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4

            try:
                writer.close()
            except Exception as save_e:
                print(f"[ERROR] 儲存 Excel 失敗: {save_e}")
            finally:
                shutil.rmtree(temp_dir, ignore_errors=True)

            return output_path